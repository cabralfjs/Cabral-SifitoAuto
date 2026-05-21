#!/usr/bin/env python3
"""
update_sifito.py
────────────────
Abre https://sifito.dgav.pt/divulgacao/usos com um browser headless (Playwright),
clica em "Exportar para Excel", converte o .xlsx para JSON compacto,
e guarda em sifito_data.json pronto a ser servido pelo site.

Dependências:
    pip install playwright openpyxl
    playwright install chromium
"""

import asyncio, json, os, sys, tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from playwright.async_api import async_playwright

SIFITO_URL  = "https://sifito.dgav.pt/divulgacao/usos"
OUTPUT_FILE = Path(__file__).parent / "sifito_data.json"
TIMEOUT_MS  = 60_000    # timeout de navegação/elementos
EXPORT_WAIT = 300       # segundos máximos a aguardar o export (5 min para 70k linhas)

COL_MAP = {
    0:  "cultura",
    3:  "sit_particular",
    4:  "ambiente",
    5:  "inimigo",
    6:  "nome_cient",
    8:  "uso_menor",
    12: "produto",
    13: "autorizacao",
    14: "numero",
    15: "funcao",
    16: "substancia",
    17: "epoca",
    18: "tecnica",
    19: "num_max_intervalo",
    20: "concentracao",
    21: "vol_calda",
    22: "dose",
    23: "intervalo_seg",
    24: "restricoes",
    26: "validade",
    27: "limite_comerc",
    28: "limite_util",
}


async def download_excel() -> bytes:
    """Abre o SIFITO e captura o ficheiro Excel com múltiplas estratégias."""
    print("🌐  A abrir o browser…")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page    = await context.new_page()

        # ── Estratégia 1: interceptar respostas de rede (export server-side) ──
        excel_from_network: list[bytes] = []

        async def on_response(response):
            ct = response.headers.get("content-type", "")
            if any(x in ct.lower() for x in
                   ["spreadsheet", "excel", "officedocument", "octet-stream"]):
                try:
                    body = await response.body()
                    # XLSX começa com assinatura ZIP: PK\x03\x04
                    if len(body) > 4 and body[:4] == b'PK\x03\x04':
                        excel_from_network.append(body)
                        print(f"📥  Excel capturado via rede ({len(body):,} bytes)")
                except Exception:
                    pass

        page.on("response", on_response)

        # ── Estratégia 2: evento de download do Playwright ──
        download_future: asyncio.Future = asyncio.get_event_loop().create_future()

        def on_download(dl):
            if not download_future.done():
                download_future.set_result(dl)

        page.on("download", on_download)

        # ── Navegar e aguardar tabela ──
        print(f"🌐  A navegar para {SIFITO_URL}")
        await page.goto(SIFITO_URL, timeout=TIMEOUT_MS, wait_until="networkidle")

        print("⏳  A aguardar carregamento da tabela…")
        await page.locator("table.k-grid-table tr").first.wait_for(timeout=TIMEOUT_MS)
        print("✅  Tabela carregada.")

        # ── Encontrar e clicar no botão (button pai, não o span) ──
        btn = page.locator(
            "button:has(span.k-button-text)",
            has_text="Exportar para Excel"
        ).first
        await btn.wait_for(timeout=TIMEOUT_MS)
        print("✅  Botão encontrado. A clicar…")

        # Usar JavaScript click para garantir que o evento dispara correctamente
        await btn.evaluate("el => el.click()")
        print(f"⏳  A aguardar geração do Excel (até {EXPORT_WAIT}s para 70k linhas)…")

        # ── Aguardar qualquer uma das estratégias ──
        for elapsed in range(EXPORT_WAIT):
            await asyncio.sleep(1)

            # Estratégia 1: rede
            if excel_from_network:
                await browser.close()
                print(f"✅  Capturado via rede ao fim de {elapsed+1}s")
                return excel_from_network[0]

            # Estratégia 2: evento download
            if download_future.done():
                dl = download_future.result()
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    await dl.save_as(tmp.name)
                    data = Path(tmp.name).read_bytes()
                    Path(tmp.name).unlink()
                await browser.close()
                print(f"✅  Capturado via evento download ao fim de {elapsed+1}s")
                return data

            # Log de progresso a cada 30s
            if elapsed > 0 and elapsed % 30 == 0:
                print(f"⏳  {elapsed}s passados, ainda a aguardar…")

        # ── Estratégia 3: chamar o export via JavaScript directamente ──
        print("⚠️  Timeout expirou. A tentar export via JavaScript…")
        try:
            await page.evaluate("""
                () => {
                    const grids = document.querySelectorAll('[data-role="grid"]');
                    if (grids.length > 0) {
                        const w = window.kendo || window.jQuery && jQuery(grids[0]).data('kendoGrid');
                        if (w && w.saveAsExcel) w.saveAsExcel();
                        else if (jQuery) jQuery(grids[0]).data('kendoGrid').saveAsExcel();
                    }
                }
            """)
            # Aguarda mais 60s
            for elapsed in range(60):
                await asyncio.sleep(1)
                if excel_from_network:
                    await browser.close()
                    return excel_from_network[0]
                if download_future.done():
                    dl = download_future.result()
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                        await dl.save_as(tmp.name)
                        data = Path(tmp.name).read_bytes()
                        Path(tmp.name).unlink()
                    await browser.close()
                    return data
        except Exception as e:
            print(f"⚠️  JavaScript export falhou: {e}")

        await browser.close()
        raise RuntimeError(
            "Não foi possível capturar o download após todas as estratégias. "
            "O SIFITO pode ter mudado de estrutura ou o export demora demasiado."
        )


def xlsx_to_records(xlsx_bytes: bytes) -> list[dict]:
    """Converte o .xlsx (em bytes) para lista de dicts."""
    print("🔄  A converter Excel → JSON…")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name

    wb   = load_workbook(tmp_path, read_only=True)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    Path(tmp_path).unlink()

    records = []
    for row in rows[2:]:   # linha 0 = cabeçalho grupo, linha 1 = colunas, linha 2+ = dados
        obj = {}
        for idx, key in COL_MAP.items():
            val = row[idx] if idx < len(row) else None
            if val is None:
                obj[key] = ""
            elif isinstance(val, datetime):
                obj[key] = val.strftime("%Y-%m-%d")
            else:
                v = str(val).strip()
                obj[key] = "" if v == "-" else v
        records.append(obj)

    print(f"✅  {len(records):,} registos convertidos")
    return records


def save_json(records: list[dict]):
    today = datetime.now().strftime("%d/%m/%Y")
    out   = {"date": today, "records": records}
    OUTPUT_FILE.write_text(
        json.dumps(out, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )
    size_mb = OUTPUT_FILE.stat().st_size / 1_048_576
    print(f"💾  Guardado em {OUTPUT_FILE}  ({size_mb:.1f} MB)")


async def main():
    print("=" * 50)
    print("  SIFITO Updater")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)
    try:
        xlsx_bytes = await download_excel()
        records    = xlsx_to_records(xlsx_bytes)
        save_json(records)
        print("\n🎉  Actualização concluída com sucesso!")
    except Exception as e:
        print(f"\n❌  Erro: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
